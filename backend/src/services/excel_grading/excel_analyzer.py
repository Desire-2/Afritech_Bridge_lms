"""
Excel File Analyzer - Core workbook structure analysis.

Inspects: sheet names, named ranges, hidden sheets, cell data,
data validation, protection, and general workbook metadata.

Uses openpyxl for .xlsx/.xlsm, pandas for .csv/.xls fallback.
"""

import logging
import json
import re
from io import BytesIO
from typing import Dict, Any, List, Optional, Tuple

logger = logging.getLogger(__name__)

# Lazy imports to avoid startup cost
_openpyxl = None
_pandas = None


def _get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        import openpyxl
        _openpyxl = openpyxl
    return _openpyxl


def _get_pandas():
    global _pandas
    if _pandas is None:
        import pandas
        _pandas = pandas
    return _pandas


class ExcelAnalyzer:
    """Analyzes Excel workbook structure, sheets, and metadata."""

    def __init__(self, file_bytes: bytes, file_name: str):
        self.file_bytes = file_bytes
        self.file_name = file_name
        self.ext = file_name.rsplit('.', 1)[-1].lower() if '.' in file_name else ''
        self.workbook = None
        self.analysis: Dict[str, Any] = {}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def analyze(self) -> Dict[str, Any]:
        """Run full workbook structure analysis and return results dict."""
        try:
            if self.ext in ('xlsx', 'xlsm'):
                return self._analyze_openpyxl()
            elif self.ext == 'xls':
                return self._analyze_xls()
            elif self.ext == 'csv':
                return self._analyze_csv()
            else:
                return {'error': f'Unsupported file type: .{self.ext}', 'sheets': []}
        except Exception as e:
            logger.error(f"Excel analysis failed for {self.file_name}: {e}")
            return {'error': str(e), 'sheets': []}

    # ------------------------------------------------------------------
    # .xlsx / .xlsm  (openpyxl)
    # ------------------------------------------------------------------

    def _analyze_openpyxl(self) -> Dict[str, Any]:
        openpyxl = _get_openpyxl()
        buf = BytesIO(self.file_bytes)

        # keep_vba=True so .xlsm macros are preserved in wb.vba_archive
        self.workbook = openpyxl.load_workbook(
            buf,
            data_only=False,   # we want formulas, not cached values
            keep_vba=True,
            keep_links=False,
            read_only=False,
        )
        wb = self.workbook

        sheets_info = []
        total_formulas = 0
        total_cells_with_data = 0

        for ws in wb.worksheets:
            sheet_data = self._analyze_sheet(ws)
            sheets_info.append(sheet_data)
            total_formulas += sheet_data.get('formula_count', 0)
            total_cells_with_data += sheet_data.get('data_cell_count', 0)

        # Named ranges
        named_ranges = []
        try:
            for defn in wb.defined_names.values():
                named_ranges.append({
                    'name': defn.name,
                    'value': str(defn.attr_text),
                    'scope': defn.localSheetId,
                })
        except Exception:
            # Fallback for different openpyxl versions
            try:
                for defn in wb.defined_names:
                    named_ranges.append({
                        'name': str(defn) if isinstance(defn, str) else getattr(defn, 'name', str(defn)),
                    })
            except Exception:
                pass

        # Charts across all sheets
        total_charts = sum(len(ws._charts) for ws in wb.worksheets)

        # Pivot caches (from workbook internals)
        pivot_cache_count = 0
        try:
            if hasattr(wb, '_pivots'):
                pivot_cache_count = len(wb._pivots)
            elif hasattr(wb, 'pivot_caches'):
                pivot_cache_count = len(wb.pivot_caches)
        except Exception:
            pass

        # Hidden sheets
        hidden_sheets = [ws.title for ws in wb.worksheets
                         if ws.sheet_state != 'visible']

        self.analysis = {
            'file_name': self.file_name,
            'file_type': self.ext,
            'sheet_count': len(wb.sheetnames),
            'sheet_names': list(wb.sheetnames),
            'hidden_sheets': hidden_sheets,
            'named_ranges': named_ranges,
            'named_range_count': len(named_ranges),
            'total_formulas': total_formulas,
            'total_data_cells': total_cells_with_data,
            'total_charts': total_charts,
            'pivot_cache_count': pivot_cache_count,
            'has_vba': wb.vba_archive is not None,
            'sheets': sheets_info,
        }
        return self.analysis

    def _analyze_sheet(self, ws) -> Dict[str, Any]:
        """Analyze a single worksheet."""
        formulas = []
        data_cell_count = 0
        cell_errors = []
        data_types_found = set()
        has_conditional_formatting = len(ws.conditional_formatting) > 0
        cf_rules_count = len(ws.conditional_formatting)

        # Data validations
        data_validations = []
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                data_validations.append({
                    'type': dv.type,
                    'formula1': str(dv.formula1) if dv.formula1 else None,
                    'ranges': str(dv.sqref),
                })

        # Merged cells
        merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]

        # Protection
        is_protected = ws.protection.sheet if ws.protection else False

        # Scan cells (limit scan for very large sheets)
        max_rows = min(ws.max_row or 0, 500)
        max_cols = min(ws.max_column or 0, 50)

        for row in ws.iter_rows(min_row=1, max_row=max_rows,
                                min_col=1, max_col=max_cols):
            for cell in row:
                if cell.value is not None:
                    data_cell_count += 1
                    data_types_found.add(type(cell.value).__name__)

                    # Check for formulas
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value,
                            'sheet': ws.title,
                        })

                    # Check for errors
                    if isinstance(cell.value, str) and cell.value in (
                        '#DIV/0!', '#N/A', '#NAME?', '#NULL!',
                        '#NUM!', '#REF!', '#VALUE!', '#GETTING_DATA'
                    ):
                        cell_errors.append({
                            'cell': cell.coordinate,
                            'error': cell.value,
                            'sheet': ws.title,
                        })

        # Charts in this sheet
        charts_info = []
        for chart in ws._charts:
            charts_info.append({
                'type': chart.__class__.__name__,
                'title': str(chart.title) if chart.title else None,
                'style': getattr(chart, 'style', None),
            })

        # Pivot tables
        pivots_info = []
        if hasattr(ws, '_pivots') and ws._pivots:
            for pt in ws._pivots:
                pivots_info.append({
                    'name': getattr(pt, 'name', 'PivotTable'),
                    'location': str(getattr(pt, 'location', '')),
                })

        return {
            'name': ws.title,
            'is_hidden': ws.sheet_state != 'visible',
            'row_count': ws.max_row or 0,
            'column_count': ws.max_column or 0,
            'data_cell_count': data_cell_count,
            'formula_count': len(formulas),
            'formulas': formulas[:100],  # cap for large files
            'cell_errors': cell_errors,
            'charts': charts_info,
            'chart_count': len(charts_info),
            'pivot_tables': pivots_info,
            'pivot_count': len(pivots_info),
            'has_conditional_formatting': has_conditional_formatting,
            'conditional_formatting_rules': cf_rules_count,
            'data_validations': data_validations,
            'merged_cells': merged_ranges[:30],
            'merged_cell_count': len(merged_ranges),
            'is_protected': is_protected,
            'data_types_found': list(data_types_found),
        }

    # ------------------------------------------------------------------
    # .xls fallback (pandas)
    # ------------------------------------------------------------------

    def _analyze_xls(self) -> Dict[str, Any]:
        """Basic analysis for legacy .xls files using pandas."""
        pd = _get_pandas()
        buf = BytesIO(self.file_bytes)
        try:
            xls = pd.ExcelFile(buf, engine='xlrd')
        except Exception:
            # Fallback to openpyxl won't work for .xls; try xlrd
            return {'error': 'xlrd not installed for .xls support', 'sheets': []}

        sheets_info = []
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name, header=None)
            sheets_info.append({
                'name': name,
                'is_hidden': False,
                'row_count': len(df),
                'column_count': len(df.columns),
                'data_cell_count': int(df.count().sum()),
                'formula_count': 0,  # pandas reads values, not formulas
                'formulas': [],
                'cell_errors': [],
                'charts': [],
                'chart_count': 0,
                'pivot_tables': [],
                'pivot_count': 0,
                'has_conditional_formatting': False,
                'conditional_formatting_rules': 0,
                'data_validations': [],
                'merged_cells': [],
                'merged_cell_count': 0,
                'is_protected': False,
                'data_types_found': list(set(str(dt) for dt in df.dtypes)),
            })

        return {
            'file_name': self.file_name,
            'file_type': self.ext,
            'sheet_count': len(xls.sheet_names),
            'sheet_names': xls.sheet_names,
            'hidden_sheets': [],
            'named_ranges': [],
            'named_range_count': 0,
            'total_formulas': 0,
            'total_data_cells': sum(s['data_cell_count'] for s in sheets_info),
            'total_charts': 0,
            'pivot_cache_count': 0,
            'has_vba': False,
            'sheets': sheets_info,
            'note': '.xls format has limited analysis capabilities; formulas not extractable via pandas.',
        }

    # ------------------------------------------------------------------
    # .csv fallback
    # ------------------------------------------------------------------

    def _analyze_csv(self) -> Dict[str, Any]:
        """Analyze CSV file using pandas."""
        pd = _get_pandas()
        buf = BytesIO(self.file_bytes)
        try:
            df = pd.read_csv(buf, header=None, nrows=1000)
        except Exception as e:
            return {'error': f'CSV parse error: {e}', 'sheets': []}

        sheet = {
            'name': 'Sheet1',
            'is_hidden': False,
            'row_count': len(df),
            'column_count': len(df.columns),
            'data_cell_count': int(df.count().sum()),
            'formula_count': 0,
            'formulas': [],
            'cell_errors': [],
            'charts': [],
            'chart_count': 0,
            'pivot_tables': [],
            'pivot_count': 0,
            'has_conditional_formatting': False,
            'conditional_formatting_rules': 0,
            'data_validations': [],
            'merged_cells': [],
            'merged_cell_count': 0,
            'is_protected': False,
            'data_types_found': list(set(str(dt) for dt in df.dtypes)),
        }

        return {
            'file_name': self.file_name,
            'file_type': 'csv',
            'sheet_count': 1,
            'sheet_names': ['Sheet1'],
            'hidden_sheets': [],
            'named_ranges': [],
            'named_range_count': 0,
            'total_formulas': 0,
            'total_data_cells': sheet['data_cell_count'],
            'total_charts': 0,
            'pivot_cache_count': 0,
            'has_vba': False,
            'sheets': [sheet],
            'note': 'CSV files have no formulas, charts, or pivot tables.',
        }
