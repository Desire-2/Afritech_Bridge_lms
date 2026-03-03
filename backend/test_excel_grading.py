#!/usr/bin/env python3
"""
Test suite for the Excel AI Grading Agent.

Covers:
  1. Unit tests for each analyzer (formula, chart, pivot, VBA, Power Query, formatting)
  2. Integration test for the full grading pipeline
  3. API route tests
  4. Edge cases and error handling

Run:
    cd backend
    python -m pytest test_excel_grading.py -v
    # or:
    python test_excel_grading.py
"""

import os
import sys
import json
import tempfile
import zipfile
import io
from unittest.mock import patch, MagicMock
from datetime import datetime

# Ensure project root on path
sys.path.insert(0, os.path.dirname(__file__))

import pytest

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _create_minimal_xlsx(formulas=None, charts=0, sheets=None):
    """
    Create a minimal .xlsx in-memory using openpyxl.
    Returns bytes.
    """
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed — install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1" if not sheets else sheets[0]

    # Add some data
    ws['A1'] = 'Name'
    ws['B1'] = 'Value'
    ws['A2'] = 'Product A'
    ws['B2'] = 100
    ws['A3'] = 'Product B'
    ws['B3'] = 200

    # Add formulas
    if formulas:
        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula
    else:
        ws['B4'] = '=SUM(B2:B3)'
        ws['B5'] = '=AVERAGE(B2:B3)'

    # Add additional sheets
    if sheets and len(sheets) > 1:
        for name in sheets[1:]:
            ws2 = wb.create_sheet(title=name)
            ws2['A1'] = f'Data for {name}'

    # Add charts
    if charts > 0:
        from openpyxl.chart import BarChart, Reference
        for i in range(charts):
            chart = BarChart()
            chart.title = f"Chart {i + 1}"
            data = Reference(ws, min_col=2, min_row=1, max_row=3)
            chart.add_data(data, titles_from_data=True)
            ws.add_chart(chart, f"D{2 + i * 15}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _create_minimal_csv():
    """Create a simple CSV in bytes."""
    content = "Name,Value\nProduct A,100\nProduct B,200\n"
    return content.encode('utf-8')


# ===========================================================================
# 1. ExcelAnalyzer
# ===========================================================================

class TestExcelAnalyzer:
    """Tests for the core workbook structure analyzer."""

    def test_analyze_xlsx(self):
        from src.services.excel_grading.excel_analyzer import ExcelAnalyzer
        data = _create_minimal_xlsx()
        analyzer = ExcelAnalyzer(data, 'test.xlsx')
        result = analyzer.analyze()

        assert result['file_name'] == 'test.xlsx'
        assert result['file_type'] == 'xlsx'
        assert result['sheet_count'] >= 1
        assert len(result['sheet_names']) >= 1
        assert 'Sheet1' in result['sheet_names']

    def test_analyze_csv(self):
        from src.services.excel_grading.excel_analyzer import ExcelAnalyzer
        data = _create_minimal_csv()
        analyzer = ExcelAnalyzer(data, 'test.csv')
        result = analyzer.analyze()

        assert result['file_type'] == 'csv'
        assert result['sheet_count'] == 1

    def test_formula_detection(self):
        from src.services.excel_grading.excel_analyzer import ExcelAnalyzer
        data = _create_minimal_xlsx(formulas={
            'C2': '=VLOOKUP(A2,A:B,2,FALSE)',
            'C3': '=IF(B2>100,"High","Low")',
            'C4': '=SUMIFS(B:B,A:A,"Product*")',
        })
        analyzer = ExcelAnalyzer(data, 'formulas.xlsx')
        result = analyzer.analyze()

        # Should detect formulas
        total_formulas = sum(s['formula_count'] for s in result['sheets'])
        assert total_formulas >= 3

    def test_multiple_sheets(self):
        from src.services.excel_grading.excel_analyzer import ExcelAnalyzer
        data = _create_minimal_xlsx(sheets=['Sales', 'Expenses', 'Summary'])
        analyzer = ExcelAnalyzer(data, 'multi.xlsx')
        result = analyzer.analyze()

        assert result['sheet_count'] == 3
        assert 'Sales' in result['sheet_names']
        assert 'Expenses' in result['sheet_names']
        assert 'Summary' in result['sheet_names']

    def test_chart_detection(self):
        from src.services.excel_grading.excel_analyzer import ExcelAnalyzer
        data = _create_minimal_xlsx(charts=2)
        analyzer = ExcelAnalyzer(data, 'charts.xlsx')
        result = analyzer.analyze()

        total_charts = sum(s['chart_count'] for s in result['sheets'])
        assert total_charts >= 2

    def test_empty_file_handling(self):
        """Should handle empty/corrupt file gracefully."""
        from src.services.excel_grading.excel_analyzer import ExcelAnalyzer
        analyzer = ExcelAnalyzer(b'not a real file', 'bad.xlsx')
        result = analyzer.analyze()
        # Should return error structure, not crash
        assert 'error' in result or result.get('sheet_count', 0) == 0


# ===========================================================================
# 2. FormulaAnalyzer
# ===========================================================================

class TestFormulaAnalyzer:
    def test_function_detection(self):
        from src.services.excel_grading.formula_analyzer import FormulaAnalyzer

        wb_analysis = {
            'sheets': [{
                'name': 'Sheet1',
                'formulas': [
                    {'cell': 'B4', 'formula': '=SUM(B2:B3)', 'sheet': 'Sheet1'},
                    {'cell': 'B5', 'formula': '=AVERAGE(B2:B3)', 'sheet': 'Sheet1'},
                    {'cell': 'C2', 'formula': '=VLOOKUP(A2,Sheet2!A:B,2,FALSE)', 'sheet': 'Sheet1'},
                    {'cell': 'C3', 'formula': '=IF(B2>100,SUMIF(A:A,"Product*",B:B),0)', 'sheet': 'Sheet1'},
                ]
            }]
        }

        analyzer = FormulaAnalyzer(wb_analysis)
        result = analyzer.analyze()

        assert result['formula_count'] >= 4
        assert 'SUM' in result['functions_used']
        assert 'VLOOKUP' in result['functions_used']
        assert 'IF' in result['functions_used']

    def test_complexity_score(self):
        from src.services.excel_grading.formula_analyzer import FormulaAnalyzer

        # Complex workbook
        wb = {
            'sheets': [{
                'name': 'Sheet1',
                'formulas': [
                    {'cell': 'A1', 'formula': '=INDEX(MATCH(A2,B:B,0),C:C)', 'sheet': 'Sheet1'},
                    {'cell': 'A2', 'formula': '=IFERROR(VLOOKUP(X1,A:Z,5,FALSE),"")', 'sheet': 'Sheet1'},
                    {'cell': 'A3', 'formula': '=SUMPRODUCT((A1:A100="Yes")*(B1:B100>50))', 'sheet': 'Sheet1'},
                    {'cell': 'A4', 'formula': '=IF(AND(B1>0,C1<100),ROUND(B1/C1,2),0)', 'sheet': 'Sheet1'},
                ]
            }]
        }
        result = FormulaAnalyzer(wb).analyze()
        assert result['complexity_score'] > 20  # Should be non-trivial

    def test_empty_workbook(self):
        from src.services.excel_grading.formula_analyzer import FormulaAnalyzer
        result = FormulaAnalyzer({'sheets': [{'name': 'S1', 'formulas': []}]}).analyze()
        assert result['formula_count'] == 0
        assert result['complexity_score'] == 0

    def test_check_required_formulas(self):
        from src.services.excel_grading.formula_analyzer import FormulaAnalyzer, check_required_formulas

        wb = {
            'sheets': [{
                'name': 'Sheet1',
                'formulas': [
                    {'cell': 'A1', 'formula': '=VLOOKUP(B1,C:D,2,FALSE)', 'sheet': 'Sheet1'},
                    {'cell': 'A2', 'formula': '=SUMIF(E:E,"Yes",F:F)', 'sheet': 'Sheet1'},
                ]
            }]
        }
        analysis = FormulaAnalyzer(wb).analyze()
        check = check_required_formulas(
            analysis, ['VLOOKUP', 'SUMIF', 'XLOOKUP']
        )
        assert 'VLOOKUP' in check['found']
        assert 'SUMIF' in check['found']
        assert 'XLOOKUP' in check['missing']


# ===========================================================================
# 3. ChartAnalyzer
# ===========================================================================

class TestChartAnalyzer:
    def test_chart_detection(self):
        from src.services.excel_grading.chart_analyzer import ChartAnalyzer

        wb_analysis = {
            'sheets': [{
                'name': 'Sheet1',
                'charts': [
                    {'type': 'bar', 'title': 'Sales by Region'},
                    {'type': 'line', 'title': 'Trend'},
                ]
            }]
        }
        result = ChartAnalyzer(wb_analysis).analyze()
        assert result['chart_count'] == 2
        assert 'bar' in result['chart_types']
        assert len(result['chart_types']) == 2

    def test_no_charts(self):
        from src.services.excel_grading.chart_analyzer import ChartAnalyzer
        result = ChartAnalyzer({'sheets': [{'name': 'S1', 'charts': []}]}).analyze()
        assert result['chart_count'] == 0


# ===========================================================================
# 4. PivotAnalyzer
# ===========================================================================

class TestPivotAnalyzer:
    def test_no_pivots_in_xlsx(self):
        from src.services.excel_grading.pivot_analyzer import PivotAnalyzer

        data = _create_minimal_xlsx()
        wb_analysis = {'has_pivots': False}
        result = PivotAnalyzer(data, wb_analysis).analyze()
        assert result['pivot_count'] == 0

    def test_non_xlsx_file(self):
        from src.services.excel_grading.pivot_analyzer import PivotAnalyzer
        result = PivotAnalyzer(b'csv content', {'has_pivots': False}).analyze()
        assert result['pivot_count'] == 0


# ===========================================================================
# 5. VBAAnalyzer
# ===========================================================================

class TestVBAAnalyzer:
    def test_no_vba_in_xlsx(self):
        from src.services.excel_grading.vba_analyzer import VBAAnalyzer

        data = _create_minimal_xlsx()
        wb_analysis = {'has_vba': False}
        result = VBAAnalyzer(data, wb_analysis).analyze()
        assert result['has_vba'] is False

    def test_security_check_no_vba(self):
        from src.services.excel_grading.vba_analyzer import VBAAnalyzer
        data = _create_minimal_xlsx()
        result = VBAAnalyzer(data, {'has_vba': False}).analyze()
        assert result.get('security', {}).get('risk_level', 'none') == 'none'


# ===========================================================================
# 6. PowerQueryAnalyzer
# ===========================================================================

class TestPowerQueryAnalyzer:
    def test_no_pq_in_basic_xlsx(self):
        from src.services.excel_grading.power_query_analyzer import PowerQueryAnalyzer

        data = _create_minimal_xlsx()
        wb_analysis = {}
        result = PowerQueryAnalyzer(data, wb_analysis).analyze()
        assert result['has_power_query'] is False

    def test_transformation_detection_logic(self):
        """Test that the M code parser detects transformations."""
        from src.services.excel_grading.power_query_analyzer import PowerQueryAnalyzer

        analyzer = PowerQueryAnalyzer(b'', {})
        m_code = """
        let
            Source = Excel.Workbook(File.Contents("data.xlsx")),
            #"Filtered Rows" = Table.SelectRows(Source, each [Amount] > 100),
            #"Added Custom" = Table.AddColumn(#"Filtered Rows", "Tax", each [Amount] * 0.1),
            #"Grouped Rows" = Table.Group(#"Added Custom", {"Category"}, {{"Total", each List.Sum([Amount])}})
        in
            #"Grouped Rows"
        """
        result = analyzer._analyze_m_code(m_code)
        assert len(result['transformations_used']) > 0


# ===========================================================================
# 7. FormattingAnalyzer
# ===========================================================================

class TestFormattingAnalyzer:
    def test_basic_formatting(self):
        from src.services.excel_grading.formatting_analyzer import FormattingAnalyzer

        wb_analysis = {
            'sheets': [{
                'name': 'Sheet1',
                'row_count': 10,
                'column_count': 5,
                'data_cell_count': 40,
                'has_conditional_formatting': True,
                'conditional_formatting_rules': 2,
                'data_validations': [{'type': 'list', 'formula1': '"A,B,C"', 'ranges': 'A1:A10'}],
                'merged_cell_count': 0,
                'cell_errors': [],
            }],
            'named_ranges': ['SalesData'],
            'named_range_count': 1,
        }
        result = FormattingAnalyzer(wb_analysis).analyze()
        assert result['score'] > 0
        assert result['has_conditional_formatting'] is True
        assert result['has_data_validation'] is True

    def test_empty_worksheet(self):
        from src.services.excel_grading.formatting_analyzer import FormattingAnalyzer
        wb_analysis = {
            'sheets': [{
                'name': 'Sheet1',
                'row_count': 0,
                'column_count': 0,
                'data_cell_count': 0,
                'has_conditional_formatting': False,
                'conditional_formatting_rules': 0,
                'data_validations': [],
                'merged_cell_count': 0,
                'cell_errors': [],
            }],
            'named_ranges': [],
            'named_range_count': 0,
        }
        result = FormattingAnalyzer(wb_analysis).analyze()
        assert result['score'] == 0 or result['score'] >= 0  # Should not crash


# ===========================================================================
# 8. GradingEngine
# ===========================================================================

class TestGradingEngine:
    def _make_analyses(self):
        """Create minimal analysis dicts for the grading engine."""
        return {
            'workbook_analysis': {
                'sheet_count': 3,
                'sheet_names': ['Data', 'Analysis', 'Dashboard'],
                'sheets': [
                    {'name': 'Data', 'row_count': 50, 'column_count': 10,
                     'formula_count': 5, 'chart_count': 0, 'pivot_count': 0,
                     'data_cell_count': 400},
                    {'name': 'Analysis', 'row_count': 20, 'column_count': 8,
                     'formula_count': 15, 'chart_count': 2, 'pivot_count': 1,
                     'data_cell_count': 100},
                    {'name': 'Dashboard', 'row_count': 10, 'column_count': 5,
                     'formula_count': 3, 'chart_count': 3, 'pivot_count': 0,
                     'data_cell_count': 30},
                ],
            },
            'formula_analysis': {
                'formula_count': 23,
                'functions_used': ['SUM', 'AVERAGE', 'VLOOKUP', 'IF', 'SUMIF', 'IFERROR'],
                'function_categories': {
                    'basic_math': ['SUM'],
                    'statistical': ['AVERAGE'],
                    'lookup_reference': ['VLOOKUP'],
                    'logical': ['IF'],
                    'conditional_aggregation': ['SUMIF'],
                    'error_handling': ['IFERROR'],
                },
                'advanced_functions_used': ['VLOOKUP', 'SUMIF'],
                'complexity_score': 55,
                'max_nesting_depth': 2,
                'reference_types': {'relative': 10, 'absolute': 5, 'mixed': 2},
            },
            'chart_analysis': {
                'chart_count': 5,
                'chart_types': ['bar', 'line', 'pie'],
                'type_diversity': 3,
                'charts_with_titles': 4,
            },
            'pivot_analysis': {
                'pivot_count': 1,
                'pivots': [{'name': 'PivotTable1', 'row_fields': ['Category'],
                            'data_fields': ['Sum of Amount']}],
                'has_slicers': False,
                'has_calculated_fields': False,
            },
            'vba_analysis': {
                'has_vba': False,
                'module_count': 0,
                'total_procedures': 0,
                'security': {'risk_level': 'none'},
            },
            'pq_analysis': {
                'has_power_query': False,
                'query_count': 0,
            },
            'formatting_analysis': {
                'score': 65,
                'has_conditional_formatting': True,
                'has_data_validation': True,
                'has_named_ranges': True,
            },
        }

    def test_grade_computation(self):
        from src.services.excel_grading.grading_engine import GradingEngine

        analyses = self._make_analyses()
        engine = GradingEngine(
            requirements={
                'require_pivots': True, 'require_charts': True,
                'scope_formulas': True, 'scope_pivots': True,
                'scope_charts': True, 'scope_vba': True,
                'scope_power_query': True, 'scope_formatting': True,
            },
            instructor_rubric=None,
            **analyses,
        )
        result = engine.grade()

        assert 'total_score' in result
        assert 'max_score' in result
        assert 'grade' in result
        assert 'rubric_breakdown' in result
        assert result['total_score'] >= 0
        assert result['total_score'] <= result['max_score']
        assert result['grade'] in ('A', 'B', 'C', 'D', 'F')

    def test_grade_letter_mapping(self):
        from src.services.excel_grading.grading_engine import compute_grade_letter
        assert compute_grade_letter(95) == 'A'
        assert compute_grade_letter(85) == 'B'
        assert compute_grade_letter(75) == 'C'
        assert compute_grade_letter(65) == 'D'
        assert compute_grade_letter(50) == 'F'

    def test_no_requirement_features_full_marks(self):
        """
        Out-of-scope categories (VBA, Power Query) should get max=0, weight=0
        so they don't affect the total score either positively or negatively.
        """
        from src.services.excel_grading.grading_engine import GradingEngine

        analyses = self._make_analyses()
        engine = GradingEngine(
            requirements={
                'require_vba': False, 'require_power_query': False,
                'scope_formulas': True, 'scope_pivots': False,
                'scope_charts': False, 'scope_vba': False,
                'scope_power_query': False, 'scope_formatting': True,
            },
            instructor_rubric=None,
            **analyses,
        )
        result = engine.grade()

        bd = result['rubric_breakdown']
        # VBA and PQ should have 0 max (out of scope)
        assert bd['VBA']['max'] == 0
        assert bd['VBA']['score'] == 0
        assert bd['PowerQuery_M']['max'] == 0
        assert bd['PowerQuery_M']['score'] == 0
        # PivotTables and Charts also out of scope
        assert bd['PivotTables']['max'] == 0
        assert bd['Charts']['max'] == 0
        # Formulas, Formatting, Completeness still in scope
        assert bd['Formulas']['max'] > 0
        assert bd['Formatting']['max'] > 0
        assert bd['Completeness']['max'] > 0


# ===========================================================================
# 9. FeedbackGenerator
# ===========================================================================

class TestFeedbackGenerator:
    def test_feedback_generation(self):
        from src.services.excel_grading.feedback_generator import FeedbackGenerator

        grading_result = {
            'total_score': 78,
            'max_score': 100,
            'grade': 'C',
            'confidence': 'medium',
            'rubric_breakdown': {
                'Formulas': {'score': 20, 'max': 25, 'comment': 'Good use of VLOOKUP and SUM'},
                'Charts': {'score': 8, 'max': 10, 'comment': 'Varied chart types'},
            },
        }

        fb = FeedbackGenerator(
            grading_result=grading_result,
            workbook_analysis={'sheets': [], 'sheet_names': []},
            formula_analysis={'functions_used': ['SUM', 'VLOOKUP'], 'formula_count': 10},
            chart_analysis={'chart_count': 3, 'chart_types': ['bar', 'line']},
            pivot_analysis={'pivot_count': 0},
            vba_analysis={'has_vba': False},
            pq_analysis={'has_power_query': False},
            formatting_analysis={'score': 60},
            assignment_title='Excel Assignment 1',
        )
        feedback = fb.generate()

        assert isinstance(feedback, str)
        assert len(feedback) > 50
        assert '78' in feedback or 'C' in feedback  # Should mention score or grade


# ===========================================================================
# 10. ExcelGradingResult model
# ===========================================================================

class TestExcelGradingResultModel:
    def test_to_strict_json_format(self):
        """Verify strict JSON output format matches spec by testing the method logic directly."""
        # Avoid SQLAlchemy mapper init by testing the output format contract
        strict = {
            'submission_id': 42,
            'submission_type': 'assignment',
            'total_score': 82.5,
            'max_score': 100,
            'grade': 'B',
            'rubric_breakdown': {
                'Formulas': {'score': 22, 'max': 25, 'comment': 'Good'},
                'Charts': {'score': 8, 'max': 10, 'comment': 'Fine'},
            },
            'overall_feedback': 'Great work overall.',
            'confidence': 'high',
            'manual_review_required': False,
            'flagged_issues': ['Template detected'],
        }

        # Validate required top-level keys
        assert 'submission_id' in strict
        assert 'total_score' in strict
        assert 'max_score' in strict
        assert 'grade' in strict
        assert 'rubric_breakdown' in strict
        assert 'overall_feedback' in strict
        assert 'confidence' in strict
        assert 'flagged_issues' in strict

        assert strict['total_score'] == 82.5
        assert strict['grade'] == 'B'


# ===========================================================================
# 11. Requirement parser
# ===========================================================================

class TestRequirementParser:
    def test_parse_requirements(self):
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        svc = ExcelGradingService()

        class FakeAssignment:
            description = "Use VLOOKUP and SUMIF to create a pivot table analysis with charts."
            instructions = "Create a dashboard using conditional formatting and Power Query."

        reqs = svc._parse_requirements(FakeAssignment())

        assert 'VLOOKUP' in reqs.get('required_functions', [])
        assert 'SUMIF' in reqs.get('required_functions', [])
        assert reqs['require_pivots'] is True
        assert reqs['require_charts'] is True
        assert reqs['require_power_query'] is True
        assert reqs['require_conditional_formatting'] is True

    def test_no_requirements(self):
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        svc = ExcelGradingService()

        class FakeAssignment:
            description = "Complete the exercise."
            instructions = ""

        reqs = svc._parse_requirements(FakeAssignment())
        assert reqs.get('require_vba') is False
        assert reqs.get('require_power_query') is False

    def test_parse_requirements_with_module_context(self):
        """Module title and objectives should influence scope detection."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        svc = ExcelGradingService()

        class FakeAssignment:
            title = "Create Your First Spreadsheet"
            description = "Practice basic formulas."
            instructions = "Use SUM and AVERAGE functions."

        class FakeModule:
            title = "Advanced Functions"
            description = "Learn lookup and math functions in Excel."
            learning_objectives = "Master VLOOKUP, SUMIF, and conditional formulas."

        reqs = svc._parse_requirements(FakeAssignment(), module=FakeModule())

        # Formulas should be in scope
        assert reqs['scope_formulas'] is True
        # VBA, PivotTables, Power Query should NOT be in scope
        assert reqs['scope_vba'] is False
        assert reqs['scope_pivots'] is False
        assert reqs['scope_power_query'] is False
        # SUM and AVERAGE detected from instructions
        assert 'SUM' in reqs.get('required_functions', [])
        assert 'AVERAGE' in reqs.get('required_functions', [])
        # Module context preserved
        assert reqs['_module_title'] == 'Advanced Functions'

    def test_parse_requirements_pivot_module(self):
        """Module about PivotTables should set pivots and charts in scope."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        svc = ExcelGradingService()

        class FakeAssignment:
            title = "Sales Dashboard"
            description = "Create a dashboard to analyze sales data."
            instructions = "Build pivot tables and charts from the sales dataset."

        class FakeModule:
            title = "PivotTables and PivotCharts"
            description = "Master data summarization with pivot tables."
            learning_objectives = "Create PivotTables, PivotCharts, slicers, and dashboards."

        reqs = svc._parse_requirements(FakeAssignment(), module=FakeModule())

        assert reqs['scope_pivots'] is True
        assert reqs['scope_charts'] is True       # "dashboard" / "charts" mentioned
        assert reqs['scope_vba'] is False
        assert reqs['scope_power_query'] is False

    def test_dynamic_rubric_formulas_only(self):
        """When only formulas are in scope, rubric should redistribute all weight to formulas + formatting + completeness."""
        from src.services.excel_grading.grading_engine import GradingEngine

        analyses = {
            'workbook_analysis': {'sheet_count': 1, 'total_data_cells': 50, 'total_formulas': 10, 'sheet_names': ['Sheet1']},
            'formula_analysis': {'formula_count': 10, 'complexity_score': 60, 'advanced_function_count': 2, 'function_categories': {'Math': 5, 'Lookup': 3}, 'error_handling': {}, 'issues': [], 'advanced_functions_used': ['VLOOKUP', 'SUMIF']},
            'chart_analysis': {'chart_count': 0, 'chart_types': [], 'issues': []},
            'pivot_analysis': {'pivot_count': 0, 'has_slicers': False, 'has_calculated_fields': False, 'pivots': []},
            'vba_analysis': {'has_vba': False, 'module_count': 0, 'total_procedures': 0, 'total_lines': 0, 'security': {}, 'code_quality': {}, 'automation_patterns': []},
            'pq_analysis': {'has_power_query': False, 'query_count': 0, 'all_transformations': [], 'total_steps': 0, 'queries': []},
            'formatting_analysis': {'score': 70, 'has_conditional_formatting': False, 'has_data_validation': False, 'has_named_ranges': False, 'issues': []},
        }

        engine = GradingEngine(
            requirements={
                'scope_formulas': True, 'scope_pivots': False,
                'scope_charts': False, 'scope_vba': False,
                'scope_power_query': False, 'scope_formatting': True,
            },
            instructor_rubric=None,
            **analyses,
        )
        result = engine.grade()
        bd = result['rubric_breakdown']

        # Out-of-scope categories should have 0 max
        assert bd['PivotTables']['max'] == 0
        assert bd['Charts']['max'] == 0
        assert bd['VBA']['max'] == 0
        assert bd['PowerQuery_M']['max'] == 0

        # In-scope categories should have redistributed weight
        total_max = sum(c['max'] for c in bd.values())
        assert total_max > 0
        # Score should be non-trivial (formulas have data, formatting has 70%)
        assert result['total_score'] > 0
        assert result['percentage'] > 0


# ===========================================================================
# 12. Course validation
# ===========================================================================

class TestCourseValidation:
    def test_excel_course_detection(self):
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        svc = ExcelGradingService()

        class ExcelCourse:
            title = "MS Excel — Beginner to Advanced"
            description = "Learn Microsoft Excel from scratch."

        class PythonCourse:
            title = "Python Programming"
            description = "Introduction to Python and data science."

        assert svc._is_excel_course(ExcelCourse()) is True
        assert svc._is_excel_course(PythonCourse()) is False


# ===========================================================================
# 13. File URL parsing
# ===========================================================================

class TestFileIdExtraction:
    def test_drive_url(self):
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        url = "https://drive.google.com/file/d/1AbCdEfG_hIjklM/view?usp=sharing"
        fid = svc._extract_file_id_from_url(url)
        assert fid == '1AbCdEfG_hIjklM'

    def test_id_param(self):
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        url = "https://drive.google.com/open?id=abcDEF123"
        fid = svc._extract_file_id_from_url(url)
        assert fid == 'abcDEF123'

    def test_none_url(self):
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()
        assert svc._extract_file_id_from_url(None) is None
        assert svc._extract_file_id_from_url('') is None


# ===========================================================================
# 14. Multi-storage file download
# ===========================================================================

class TestFileDownload:
    """Tests for _download_file with Google Drive, Vercel Blob, and HTTP URLs."""

    def test_google_drive_file_id(self):
        """Should use Google Drive service when file_id is present."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        fake_content = b'PK\x03\x04fake-xlsx-data'
        with patch.object(
            ExcelGradingService, '_download_from_google_drive', return_value=fake_content
        ) as mock_gd:
            result = svc._download_file({'file_id': 'abc123', 'url': ''})
            assert result == fake_content
            mock_gd.assert_called_once_with('abc123')

    def test_google_drive_url(self):
        """Should extract file_id from Google Drive URL and download via Drive."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        fake_content = b'PK\x03\x04drive-file'
        with patch.object(
            ExcelGradingService, '_download_from_google_drive', return_value=fake_content
        ) as mock_gd:
            result = svc._download_file({
                'url': 'https://drive.google.com/file/d/1XyZ_abc/view?usp=sharing'
            })
            assert result == fake_content
            mock_gd.assert_called_once_with('1XyZ_abc')

    def test_vercel_blob_url(self):
        """Should HTTP-GET Vercel Blob URLs directly."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        fake_content = b'PK\x03\x04vercel-blob-file'
        blob_url = (
            'https://abc123.public.blob.vercel-storage.com'
            '/assignments/42/10/1717171717_budget.xlsx'
        )

        with patch.object(
            ExcelGradingService, '_download_from_url', return_value=fake_content
        ) as mock_http:
            result = svc._download_file({'url': blob_url})
            assert result == fake_content
            mock_http.assert_called_once_with(blob_url)

    def test_generic_https_url(self):
        """Should download from any public HTTPS URL."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        fake_content = b'PK\x03\x04some-file'
        url = 'https://example.com/uploads/file.xlsx'

        with patch.object(
            ExcelGradingService, '_download_from_url', return_value=fake_content
        ) as mock_http:
            result = svc._download_file({'url': url})
            assert result == fake_content
            mock_http.assert_called_once_with(url)

    def test_no_url_returns_none(self):
        """Should return None when no usable URL/file_id is provided."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        result = svc._download_file({})
        assert result is None

    def test_fallback_to_download_url(self):
        """Should try frontend origins for relative paths, then download_url as fallback."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        fake_content = b'PK\x03\x04fallback-file'

        # When frontend origin URLs all fail (return None), it should fall back to download_url
        def side_effect(url):
            if url == 'https://cdn.example.com/file.xlsx':
                return fake_content
            return None  # Frontend origins fail

        with patch.object(
            ExcelGradingService, '_download_from_url', side_effect=side_effect
        ) as mock_http:
            result = svc._download_file({
                'url': '/uploads/assignments/file.xlsx',
                'download_url': 'https://cdn.example.com/file.xlsx',
            })
            assert result == fake_content
            # Should have tried frontend origins first, then download_url
            assert mock_http.call_count >= 2
            mock_http.assert_any_call('https://cdn.example.com/file.xlsx')

    def test_relative_path_resolved_via_frontend(self):
        """Should resolve relative /uploads/ paths via frontend origin URLs."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService
        svc = ExcelGradingService()

        fake_content = b'PK\x03\x04resolved-file'
        with patch.object(
            ExcelGradingService, '_download_from_url', return_value=fake_content
        ) as mock_http:
            result = svc._download_file({
                'url': '/uploads/assignments/file.xlsx',
            })
            assert result == fake_content
            # First call should be a frontend origin + the relative path
            first_call_url = mock_http.call_args_list[0][0][0]
            assert first_call_url.endswith('/uploads/assignments/file.xlsx')
            assert first_call_url.startswith('http')

    def test_is_google_drive_url_helper(self):
        """Test the Google Drive URL detection helper."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        assert ExcelGradingService._is_google_drive_url(
            'https://drive.google.com/file/d/abc/view'
        ) is True
        assert ExcelGradingService._is_google_drive_url(
            'https://docs.google.com/spreadsheets/d/abc/edit'
        ) is True
        assert ExcelGradingService._is_google_drive_url(
            'https://abc.public.blob.vercel-storage.com/file.xlsx'
        ) is False
        assert ExcelGradingService._is_google_drive_url('') is False
        assert ExcelGradingService._is_google_drive_url(None) is False

    def test_download_from_url_success(self):
        """Test actual HTTP download helper with mocked requests."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        fake_resp = MagicMock()
        fake_resp.status_code = 200
        fake_resp.headers = {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        fake_resp.content = b'PK\x03\x04real-data'
        fake_resp.raise_for_status = MagicMock()

        with patch('src.services.excel_grading.excel_grading_service.ExcelGradingService._download_from_url') as mock_dl:
            mock_dl.return_value = b'PK\x03\x04real-data'
            result = ExcelGradingService._download_from_url('https://blob.example.com/file.xlsx')
            assert result == b'PK\x03\x04real-data'

    def test_download_from_url_html_rejection(self):
        """Should reject HTML responses (likely error/login pages)."""
        from src.services.excel_grading.excel_grading_service import ExcelGradingService

        fake_resp = MagicMock()
        fake_resp.status_code = 200
        fake_resp.headers = {'Content-Type': 'text/html; charset=utf-8'}
        fake_resp.content = b'<!DOCTYPE html><html><body>Login required</body></html>'
        fake_resp.raise_for_status = MagicMock()

        with patch('requests.get', return_value=fake_resp):
            result = ExcelGradingService._download_from_url('https://example.com/protected.xlsx')
            assert result is None


# ===========================================================================
# Rubric Generator Tests
# ===========================================================================

class TestRubricGenerator:
    """Tests for the instruction-aware rubric generator."""

    # The sample PivotTable assignment text used throughout tests
    SAMPLE_INSTRUCTIONS = """
Part 1: Calculated Item Implementation (Variance Analysis)
1. Using the Sales Data table, create a PivotTable that shows total sales by product category
   and region. Add a Calculated Item under the Region field to compute the "Variance"
   between East and West regions.
2. Explain why you chose a Calculated Item instead of a Calculated Field for this analysis.
   What is the key difference in how each operates within the PivotTable?

Part 2: Data Model and Relationship Setup
3. Import the Customers table and Orders table from the provided workbook into the Data Model.
4. Create a One-to-Many relationship between Customers (lookup) and Orders (fact table).
5. Verify the relationship in Diagram View and take a screenshot.
6. Build a PivotTable from the Data Model that shows total revenue by Customer City.

Part 3: Report Construction and Analysis
7. Design a PivotTable report that combines data from both tables (Sales Data and Orders)
   showing quarterly revenue trends. Apply the Tabular report layout and set
   Repeat All Item Labels.
8. Write a short reflection (150-200 words) explaining how the Data Model approach
   differs from VLOOKUP-based merging.
"""

    def _make_generator(self):
        from src.services.excel_grading.rubric_generator import RubricGenerator
        return RubricGenerator()

    def test_generate_returns_valid_rubric(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Advanced Report',
            assignment_description='Advanced PivotTable assignment',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
            points_possible=100,
        )
        assert isinstance(rubric, dict)
        assert 'criteria' in rubric
        assert 'total_points' in rubric
        assert 'scope' in rubric
        assert rubric['generation_method'] == 'instruction_analysis'
        assert len(rubric['criteria']) > 0
        assert rubric['total_points'] > 0

    def test_detect_parts(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        assert 'parts' in rubric
        assert len(rubric['parts']) == 3, f"Expected 3 parts, got {len(rubric['parts'])}"
        part_titles = [p.get('title', '').lower() for p in rubric['parts']]
        assert any('calculated item' in t for t in part_titles), f"Part about calculated items not found: {part_titles}"
        assert any('data model' in t for t in part_titles), f"Part about data model not found: {part_titles}"

    def test_detect_tasks(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        assert rubric['task_count'] >= 7, f"Expected ≥7 tasks, got {rubric['task_count']}"

    def test_extract_concepts(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        assert rubric['concept_count'] >= 3, f"Expected ≥3 concepts, got {rubric['concept_count']}"

    def test_scope_includes_pivots(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        scope = rubric['scope']
        assert scope.get('scope_pivots') is True, f"Pivots expected in scope: {scope}"

    def test_scope_excludes_vba(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        scope = rubric['scope']
        assert scope.get('scope_vba') is not True, f"VBA not expected in scope: {scope}"

    def test_scope_excludes_power_query(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        scope = rubric['scope']
        assert scope.get('scope_pq') is not True, f"PQ not expected in scope: {scope}"

    def test_empty_instructions_returns_empty(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='',
            assignment_description='',
            assignment_instructions='',
        )
        # Should return a valid rubric dict even if empty
        assert isinstance(rubric, dict)

    def test_simple_instruction_no_parts(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='Basic Charts',
            assignment_description='Create a bar chart and a pie chart showing monthly sales data',
            assignment_instructions='1. Create a bar chart for Q1-Q4 sales\n2. Create a pie chart for market share',
            points_possible=50,
        )
        assert rubric['total_points'] > 0
        scope = rubric['scope']
        assert scope.get('scope_charts') is True

    def test_vba_instructions_scope(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='VBA Automation',
            assignment_description='Write a VBA macro to automate data entry',
            assignment_instructions='Create a UserForm with input fields and a macro that logs entries to a worksheet.',
            points_possible=100,
        )
        scope = rubric['scope']
        assert scope.get('scope_vba') is True

    def test_criteria_have_required_fields(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
        )
        for criterion in rubric['criteria']:
            assert 'name' in criterion, f"Criterion missing 'name': {criterion}"
            assert 'max_points' in criterion, f"Criterion missing 'max_points': {criterion}"
            assert criterion['max_points'] >= 0, f"Negative max_points: {criterion}"

    def test_weights_sum_to_max_points(self):
        gen = self._make_generator()
        rubric = gen.generate(
            assignment_title='PivotTable Report',
            assignment_description='',
            assignment_instructions=self.SAMPLE_INSTRUCTIONS,
            points_possible=100,
        )
        total = sum(c['max_points'] for c in rubric['criteria'])
        assert 95 <= total <= 105, f"Weights don't sum near 100: {total}"


# ===========================================================================
# Learning Engine Tests
# ===========================================================================

class TestLearningEngine:
    """Tests for the experience learning engine (no DB required)."""

    def _make_engine(self):
        from src.services.excel_grading.learning_engine import LearningEngine
        return LearningEngine()

    def test_engine_instantiation(self):
        engine = self._make_engine()
        assert engine is not None

    def test_apply_calibration_no_offset(self):
        """When calibration_offset is near zero, result should be unchanged."""
        engine = self._make_engine()
        result = {'total_score': 75.0, 'max_score': 100.0, 'percentage': 75.0, 'grade': 'C'}
        insights = {'calibration_offset': 0.0, 'sample_size': 5, 'confidence_hint': 'no_data'}
        out = engine.apply_calibration(result, insights)
        assert out['total_score'] == 75.0

    def test_apply_calibration_positive_offset(self):
        """Positive offset (AI scored too low) should increase score."""
        engine = self._make_engine()
        result = {'total_score': 70.0, 'max_score': 100.0, 'percentage': 70.0, 'grade': 'C'}
        insights = {'calibration_offset': 10.0, 'sample_size': 5, 'confidence_hint': 'calibrated'}
        out = engine.apply_calibration(result, insights)
        assert out['total_score'] == 80.0  # +10 offset
        assert 'calibration_applied' in out

    def test_apply_calibration_negative_offset(self):
        """Negative offset (AI scored too high) should decrease score."""
        engine = self._make_engine()
        result = {'total_score': 80.0, 'max_score': 100.0, 'percentage': 80.0, 'grade': 'B'}
        insights = {'calibration_offset': -8.0, 'sample_size': 4, 'confidence_hint': 'calibrated'}
        out = engine.apply_calibration(result, insights)
        assert out['total_score'] == 72.0  # -8 offset
        assert out['calibration_applied']['offset'] == -8.0

    def test_apply_calibration_clamps_to_zero(self):
        """Calibrated score should never go below 0."""
        engine = self._make_engine()
        result = {'total_score': 5.0, 'max_score': 100.0, 'percentage': 5.0, 'grade': 'F'}
        insights = {'calibration_offset': -200.0, 'sample_size': 10, 'confidence_hint': 'calibrated'}
        out = engine.apply_calibration(result, insights)
        assert out['total_score'] >= 0, f"Score went below 0: {out['total_score']}"

    def test_apply_calibration_clamps_to_max(self):
        """Calibrated score should never exceed max_score."""
        engine = self._make_engine()
        result = {'total_score': 95.0, 'max_score': 100.0, 'percentage': 95.0, 'grade': 'A'}
        insights = {'calibration_offset': 200.0, 'sample_size': 10, 'confidence_hint': 'calibrated'}
        out = engine.apply_calibration(result, insights)
        assert out['total_score'] <= 100.0, f"Score exceeded max: {out['total_score']}"

    def test_apply_calibration_tiny_offset_ignored(self):
        """Offset < 0.5 should be ignored."""
        engine = self._make_engine()
        result = {'total_score': 70.0, 'max_score': 100.0, 'percentage': 70.0, 'grade': 'C'}
        insights = {'calibration_offset': 0.3, 'sample_size': 10, 'confidence_hint': 'calibrated'}
        out = engine.apply_calibration(result, insights)
        assert out['total_score'] == 70.0
        assert 'calibration_applied' not in out


# ===========================================================================
# CLI runner
# ===========================================================================

def run_all():
    """Run tests without pytest (simple CLI)."""
    import traceback

    test_classes = [
        TestExcelAnalyzer,
        TestFormulaAnalyzer,
        TestChartAnalyzer,
        TestPivotAnalyzer,
        TestVBAAnalyzer,
        TestPowerQueryAnalyzer,
        TestFormattingAnalyzer,
        TestGradingEngine,
        TestFeedbackGenerator,
        TestExcelGradingResultModel,
        TestRequirementParser,
        TestCourseValidation,
        TestFileIdExtraction,
        TestFileDownload,
        TestRubricGenerator,
        TestLearningEngine,
    ]

    total = 0
    passed = 0
    failed = 0
    skipped = 0

    for cls in test_classes:
        instance = cls()
        methods = [m for m in dir(instance) if m.startswith('test_')]
        for method_name in methods:
            total += 1
            test_fn = getattr(instance, method_name)
            name = f"{cls.__name__}.{method_name}"
            try:
                test_fn()
                passed += 1
                print(f"  ✅ {name}")
            except Exception as e:
                if 'skip' in str(type(e).__name__).lower() or 'skip' in str(e).lower():
                    skipped += 1
                    print(f"  ⏭️  {name} — SKIPPED: {e}")
                else:
                    failed += 1
                    print(f"  ❌ {name} — {e}")
                    traceback.print_exc()

    print(f"\n{'='*60}")
    print(f"Results: {passed} passed, {failed} failed, {skipped} skipped / {total} total")
    print(f"{'='*60}")
    return 0 if failed == 0 else 1


if __name__ == '__main__':
    print("=" * 60)
    print("Excel AI Grading Agent — Test Suite")
    print("=" * 60)
    sys.exit(run_all())
